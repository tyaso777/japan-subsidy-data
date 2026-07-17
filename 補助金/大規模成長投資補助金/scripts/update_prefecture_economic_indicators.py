#!/usr/bin/env python3
"""Build prefecture economic reference data and attach it to every case."""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from statistics import mean

import pdfplumber
from openpyxl import load_workbook


PROJECT = Path(__file__).resolve().parent.parent
RAW = PROJECT / "data" / "reference" / "raw" / "prefecture_economy"
REFERENCE = PROJECT / "data" / "reference" / "prefecture_economic_indicators.csv"
SOURCES = PROJECT / "data" / "reference" / "prefecture_economic_indicator_sources.csv"
CASES_CSV = PROJECT / "data" / "processed" / "cases.csv"
CASES_JSON = PROJECT / "html" / "data" / "cases.json"

PREFECTURES = [
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
    "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
    "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県", "岐阜県",
    "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県",
    "奈良県", "和歌山県", "鳥取県", "島根県", "岡山県", "広島県", "山口県",
    "徳島県", "香川県", "愛媛県", "高知県", "福岡県", "佐賀県", "長崎県",
    "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県",
]

REGIONS = {
    "北海道": "北海道",
    **{p: "東北" for p in PREFECTURES[1:7]},
    **{p: "関東" for p in PREFECTURES[7:14]},
    **{p: "中部" for p in PREFECTURES[14:24]},
    **{p: "近畿" for p in PREFECTURES[24:30]},
    **{p: "中国" for p in PREFECTURES[30:35]},
    **{p: "四国" for p in PREFECTURES[35:39]},
    **{p: "九州・沖縄" for p in PREFECTURES[39:]},
}

METRICS = [
    ("nominal_gpp_oku_yen", "県内総生産（名目）", "億円"),
    ("per_capita_income_thousand_yen", "1人当たり県民所得", "千円/人"),
    ("average_wage_thousand_yen_per_month", "平均賃金", "千円/月"),
    ("population_thousand", "人口", "千人"),
    ("population_change_pct", "人口増減率", "%"),
]


def prefecture_table(path: Path, column: int, divisor: float = 1.0) -> dict[str, float]:
    ws = load_workbook(path, data_only=True, read_only=True)["実数"]
    result = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        if row[1] in PREFECTURES and isinstance(row[column], (int, float)):
            result[row[1]] = round(float(row[column]) / divisor, 3)
    return result


def wage_table(path: Path) -> dict[str, float]:
    ws = load_workbook(path, data_only=False, read_only=False)["第８図"]
    series = ws._charts[0].ser[0]
    points = {point.idx: float(point.v) for point in series.val.numLit.pt}
    return {prefecture: points[index] for index, prefecture in enumerate(PREFECTURES, start=2)}


def population_table(path: Path) -> dict[str, float]:
    ws = load_workbook(path, data_only=True, read_only=True)["第2表"]
    result = {}
    for row in ws.iter_rows(min_row=13, values_only=True):
        prefecture = str(row[2] or "").strip()
        if prefecture in PREFECTURES and isinstance(row[4], (int, float)):
            result[prefecture] = float(row[4])
    return result


def population_change_table(path: Path) -> dict[str, float]:
    with pdfplumber.open(path) as pdf:
        text = pdf.pages[17].extract_text() or ""
    result = {}
    for line in text.splitlines():
        compact = re.sub(r"\s+", "", line)
        for prefecture in PREFECTURES:
            match = re.search(re.escape(prefecture) + r"(-?\d+\.\d+)(-?\d+\.\d+)", compact)
            if match:
                result[prefecture] = float(match.group(1))
    return result


def source_rows() -> list[dict[str, str]]:
    retrieved = "2026-07-17"
    return [
        {
            "indicator_key": "nominal_gpp_oku_yen", "label": "県内総生産（名目）", "unit": "億円",
            "reference_period": "2022年度", "source_agency": "内閣府 経済社会総合研究所",
            "source_title": "県民経済計算 1.県内総生産（生産側、名目）",
            "definition": "県内の経済活動によって生み出された付加価値の名目額。原表の100万円を億円へ換算。",
            "source_page_url": "https://www.esri.cao.go.jp/jp/sna/data/data_list/kenmin/files/contents/main_2022.html",
            "source_data_url": "https://www.esri.cao.go.jp/jp/sna/data/data_list/kenmin/files/contents/tables/2022/soukatu1.xlsx",
            "retrieved_at": retrieved,
            "notes": "47都道府県が同一年度でそろう最新の内閣府集約値。2008SNA・2015年基準。",
        },
        {
            "indicator_key": "per_capita_income_thousand_yen", "label": "1人当たり県民所得", "unit": "千円/人",
            "reference_period": "2022年度", "source_agency": "内閣府 経済社会総合研究所",
            "source_title": "県民経済計算 7.1人当たり県民所得",
            "definition": "県民所得を総人口で除した値。個人の給与・手取り所得そのものではない。",
            "source_page_url": "https://www.esri.cao.go.jp/jp/sna/data/data_list/kenmin/files/contents/main_2022.html",
            "source_data_url": "https://www.esri.cao.go.jp/jp/sna/data/data_list/kenmin/files/contents/tables/2022/soukatu7.xlsx",
            "retrieved_at": retrieved,
            "notes": "47都道府県が同一年度でそろう最新の内閣府集約値。",
        },
        {
            "indicator_key": "average_wage_thousand_yen_per_month", "label": "平均賃金", "unit": "千円/月",
            "reference_period": "2025年6月", "source_agency": "厚生労働省",
            "source_title": "令和7年賃金構造基本統計調査 第8図 都道府県別賃金（男女計）",
            "definition": "一般労働者・男女計の2025年6月分の所定内給与額平均。超過労働給与額を除き、税等控除前。",
            "source_page_url": "https://www.mhlw.go.jp/toukei/itiran/roudou/chingin/kouzou/z2025/index.html",
            "source_data_url": "https://www.mhlw.go.jp/toukei/itiran/roudou/chingin/kouzou/z2025/xls/zuhyo.xlsx",
            "retrieved_at": retrieved,
            "notes": "短時間労働者の1時間当たり賃金ではない。概況でいう『賃金』の定義を採用。",
        },
        {
            "indicator_key": "population_thousand", "label": "人口", "unit": "千人",
            "reference_period": "2024年10月1日", "source_agency": "総務省統計局",
            "source_title": "人口推計 第2表 都道府県、男女別人口及び人口性比",
            "definition": "2024年10月1日現在の総人口（男女計）。",
            "source_page_url": "https://www.stat.go.jp/data/jinsui/2024np/",
            "source_data_url": "https://www.stat.go.jp/data/jinsui/2024np/zuhyou/05k2024-2.xlsx",
            "retrieved_at": retrieved,
            "notes": "単位は千人。日本人人口ではなく総人口。",
        },
        {
            "indicator_key": "population_change_pct", "label": "人口増減率", "unit": "%",
            "reference_period": "2023年10月1日→2024年10月1日", "source_agency": "総務省統計局",
            "source_title": "人口推計 結果の概要 表6 都道府県別人口増減率",
            "definition": "前年10月から当年9月までの人口増減を前年10月1日人口で除した公式公表率。",
            "source_page_url": "https://www.stat.go.jp/data/jinsui/2024np/",
            "source_data_url": "https://www.stat.go.jp/data/jinsui/2024np/pdf/2024np.pdf",
            "retrieved_at": retrieved,
            "notes": "人口増減は自然増減と社会増減の合計。PDF 8ページの公表率を採用。",
        },
    ]


def split_prefectures(value: object) -> list[str]:
    values = [item.strip() for item in str(value or "").split("|")]
    return list(dict.fromkeys(item for item in values if item in PREFECTURES))


def case_values(row: dict, lookup: dict[str, dict]) -> dict:
    head = str(row.get("head_office_prefecture") or "").strip()
    projects = split_prefectures(row.get("project_location_prefectures"))
    project_regions = list(dict.fromkeys(REGIONS[p] for p in projects))
    values: dict[str, object] = {
        "head_office_region": REGIONS.get(head),
        "project_regions": " | ".join(project_regions) or None,
        "project_region_class": project_regions[0] if len(project_regions) == 1 else ("複数地方" if project_regions else None),
        "project_economic_prefecture_count": len(projects),
    }
    for key, _, _ in METRICS:
        head_value = lookup.get(head, {}).get(key)
        project_values = [lookup[p][key] for p in projects if p in lookup and lookup[p].get(key) is not None]
        values[f"head_office_{key}"] = head_value
        values[f"project_{key}_mean"] = round(mean(project_values), 3) if project_values else None
    return values


def write_csv(rows: list[dict], path: Path, fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    series = {
        "nominal_gpp_oku_yen": prefecture_table(RAW / "soukatu1_2022.xlsx", 14, 100),
        "per_capita_income_thousand_yen": prefecture_table(RAW / "soukatu7_2022.xlsx", 14),
        "average_wage_thousand_yen_per_month": wage_table(RAW / "wage_2025.xlsx"),
        "population_thousand": population_table(RAW / "population_2024.xlsx"),
        "population_change_pct": population_change_table(RAW / "population_2024_overview.pdf"),
    }
    for key, values in series.items():
        missing = sorted(set(PREFECTURES) - set(values))
        if missing:
            raise ValueError(f"{key}: missing prefectures: {missing}")
    reference_rows = [
        {"prefecture_code": f"{index:02d}", "prefecture": prefecture, "region": REGIONS[prefecture],
         **{key: series[key][prefecture] for key, _, _ in METRICS}}
        for index, prefecture in enumerate(PREFECTURES, start=1)
    ]
    write_csv(reference_rows, REFERENCE, list(reference_rows[0]))
    sources = source_rows()
    write_csv(sources, SOURCES, list(sources[0]))
    lookup = {row["prefecture"]: row for row in reference_rows}

    with CASES_CSV.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        csv_rows = list(reader)
        fields = list(reader.fieldnames or [])
    additions = list(case_values(csv_rows[0], lookup))
    fields = [field for field in fields if field not in additions] + additions
    for row in csv_rows:
        row.update({key: "" if value is None else value for key, value in case_values(row, lookup).items()})
    write_csv(csv_rows, CASES_CSV, fields)

    json_rows = json.loads(CASES_JSON.read_text(encoding="utf-8"))
    for row in json_rows:
        row.update(case_values(row, lookup))
    CASES_JSON.write_text(json.dumps(json_rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"wrote {REFERENCE} ({len(reference_rows)} prefectures)")
    print(f"updated {len(csv_rows)} CSV and {len(json_rows)} JSON cases")


if __name__ == "__main__":
    main()
